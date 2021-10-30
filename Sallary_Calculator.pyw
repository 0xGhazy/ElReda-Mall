
# A simple GUI python application to calculate the net sallary.
# Coded for elredamall.com - it's free and open source :)
# Coded by: Hossam hamdy (0xGhazy).

import os
import sys
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5 import QtWidgets, uic
from win10toast import ToastNotifier
from openpyxl import Workbook
from openpyxl.workbook import Workbook
from datetime import date
import xlsxwriter


def win10_display_notification(title, message):
    toaster = ToastNotifier()
    toaster.show_toast(title, message, duration = 3)


class SallaryReport(QtWidgets.QWidget):
    """a class that create a new window for reporting the net sallary for the employee """

    def __init__(self, name: str, sallary: float, rewards: float = 0, addational: float = 0,
                 regularity_value: float = 0, delay_hours: int = 0, absence_days: int = 0,
                 borrow_value: float = 0, deduction_value: float = 0, services_value: float = 0) -> None:

        super(QtWidgets.QWidget, self).__init__()
        os.chdir(os.path.dirname(__file__))
        uic.loadUi(r"design\report_interface.ui", self)
        self.show()

        # handling empty records
        if rewards == "":
            rewards = 0
        if addational == "":
            addational = 0
        if regularity_value == "":
            regularity_value = 0
        if delay_hours == "":
            delay_hours = 0
        if absence_days == "":
            absence_days = 0
        if borrow_value == "":
            borrow_value = 0
        if deduction_value == "":
            deduction_value = 0
        if services_value == "":
            services_value = 0

        # set values to the lbl in the window
        self.employee_name_lbl.setText(name)
        self.employee_sallary_lbl.setText(sallary)
        self.employee_rewards_lbl.setText(str(rewards))
        self.employee_additional_lbl.setText(str(addational))
        self.employee_regularity_lbl.setText(str(regularity_value))
        total_employee_rights = float(rewards) + float(addational) + float(regularity_value)
        self.total_eployee_rights.setText(str(total_employee_rights))
        self.employee_delay_lbl.setText(str(delay_hours))
        self.employee_absence_lbl.setText(str(absence_days))
        self.employee_borrow_lbl.setText(str(borrow_value))
        self.employee_deduction_lbl.setText(str(deduction_value))
        self.employee_services_lbl.setText(str(services_value))
        # some of important calculations.
        pay_per_day = (float(sallary) / 30)
        pay_per_hour = pay_per_day / 9
        get_delay = pay_per_hour * int(delay_hours)
        get_absence = pay_per_day * int(absence_days)
        get_borrow = float(borrow_value)
        get_deduction = float(deduction_value)
        get_service = float(services_value)
        employee_loan = get_delay + get_absence + get_borrow + get_deduction + get_service
        self.total_employee_loan_lbl.setText(str("{:.2f}".format(employee_loan)))
        # formating the net sallary and display it :)
        net_sallary = (float(sallary) + float(total_employee_rights)) - float(employee_loan)
        self.employee_net_sallary.setText(str("{:.2f}".format(net_sallary)))



class ApplicationUI(QtWidgets.QMainWindow):
    """ Main application Window """

    def __init__(self: "ApplicationUI") -> None:
        super(ApplicationUI, self).__init__()
        os.chdir(os.path.dirname(__file__))
        uic.loadUi(r"design\UI.ui", self)           # loading .ui design file.
        self.show()                                 # display the GUI.
        self.handleButtons()                        # loading the buttons events.


    def handleButtons(self: "ApplicationUI") -> None:
        """ Handling all buttons in the application """
        self.generate_report.clicked.connect(self.generateReportBTN)
        self.refresh_btn.clicked.connect(self.refresh)
        self.generate_finall_report.clicked.connect(self.generatFinallReport)


    def generateReportBTN(self: "ApplicationUI") -> None:
        """ Function to get a new employee record for a day """
        # Getting user inputs
        name = self.employee_name.text() # -> (+)
        sallary = self.employee_sallary.text()
        rewards = self.rewards.text()
        addational = self.additional_sallary.text()
        regularity_value = self.regularity.text() # -> (-)
        delay_hours = self.delay_hours.text()
        absence_days = self.absence_days.text()
        borrow_value = self.borrow.text()
        deduction_value = self.deduction.text()
        deduction_reason = self.deduction_reason.text()
        services_value = self.services.text()

        if  len(name) == 0 or len(sallary) == 0:
            win10_display_notification("[-] Error Message",
            "لا تستطيع ترك حقل الاسم او المرتب فارغ.")
            exit()
        else:
            self.report = SallaryReport(name, sallary, rewards, addational, regularity_value, delay_hours, \
                                        absence_days, borrow_value, deduction_value, services_value)
            
            try:
                record_content = f"{str(name)},{str(sallary)},{str(rewards)},{str(addational)},{str(regularity_value)},{str(delay_hours)},{str(absence_days)},{str(borrow_value)},{str(deduction_value)},{str(services_value)}\n"
                with open(r"{0}\data center\records.txt".format(os.getcwd()), "a+") as record:
                    record.write(record_content)
            except Exception as error_message:
                print(f"[-] \n{error_message}\n")

            # data_2_xlsx = f"{name}, {sallary}, {rewards}, {addational}, {regularity_value}, {delay_hours}, {absence_days}, {borrow_value}, {deduction_value}, {services_value}"
            # wb = Workbook()
            # ws = wb.active
            

            # dir_content = os.listdir()
            # for item in dir_content:
            #     if item != f"{date.today()}.xlsx":
            #         pass
            #     ws.append(data_2_xlsx)
            #     wb.save('name.xlsx')


    def refresh(self: "ApplicationUI") -> None:
        """ Function to clear all text records """
        self.employee_name.clear()
        self.employee_sallary.clear()
        self.rewards.clear()
        self.additional_sallary.clear()
        self.regularity.clear()
        self.delay_hours.clear()
        self.absence_days.clear()
        self.borrow.clear()
        self.deduction.clear()
        self.deduction_reason.clear()
        self.services.clear()


    def generatFinallReport(self) -> None:
        pass
        
        # reading the csv file.
        with open (r"{0}\data center\records.txt".format(os.getcwd()), "r") as csv_file:
            csv_content = csv_file.readlines()
        

        # sheet headers
        headers = [
            'اسم الموظف',
            'الراتب الاساسي',
            'المكافأت',
            'الراتب الأضافي',
            'انتظام',
            'التأخير (عدد الساعات)',
            'الغياب',
            'السلف',
            'الخصم',
            'سبب الخصم',
            'التأمينات',
            'اجمالي المستحقات',
            'اجمالي المقتطعات',
            'الراتب الصافي',
            ]


        file_name = f'Report-{date.today()}.xlsx'
        # Create a workbook
        workbook = xlsxwriter.Workbook(file_name)
        workbook_name = r"{0}\data center\{1}".format(os.getcwd(), file_name)
        wb = Workbook()
        page = wb.active
        page.title = 'Sheet1'
        # write the headers to the first line
        page.append(headers)


        # resolving csv lines content
        # 
        for line in csv_content:
            finall_quiry = []
            main_sallary = 0
            net_sallary = 0
            total_cut = 0
            total_add = 0
            get_delay = 0
            get_absence = 0
            pay_per_day = 0
            pay_per_hour = 0
            line = line.strip("\n").split(",")


            """ name, sallary, rewards, addational, regularity_value, delay_hours, absence_days,
            borrow_value, deduction_value, services_value"""

        # handling empty records


            if line[2] == "":
                line[2] = 0
            if line[3] == "":
                line[3] = 0
            if line[4] == "":
                line[4] = 0
            if line[5] == "":
                line[5] = 0
            if line[6] == "":
                line[6] = 0
            if line[7] == "":
                line[7] = 0
            if line[8] == "":
                line[8] = 0
            if line[9] == "":
                line[9] = 0


            finall_quiry.append(line[0])            # name
            finall_quiry.append(float(line[1]))     # main sallary
            finall_quiry.append(float(line[2]))     # rewards
            finall_quiry.append(float(line[3]))     # addational sallary
            finall_quiry.append(float(line[4]))     # regularity_value
            finall_quiry.append(int(line[5]))       # delay_hours
            finall_quiry.append(int(line[6]))       # absence_days
            finall_quiry.append(float(line[7]))     # borrow_value
            finall_quiry.append(float(line[8]))     # deduction_value
            finall_quiry.append(float(line[9]))     # services_value

            main_sallary = float(line[1])
            pay_per_day = (main_sallary / 30)
            pay_per_hour = pay_per_day / 9
            get_delay = pay_per_hour * int(line[5])
            get_absence = pay_per_day * int(line[6])

            total_add = float(line[2]) + float(line[3]) + float(line[3]) + float(line[4])
            finall_quiry.append(float(total_add))    # total addational

            total_cut = get_delay + get_absence + float(line[7]) + float(line[8]) + float(line[9])

            finall_quiry.append(total_add)    # total add
            finall_quiry.append(total_cut)    # total cut
            finall_quiry.append((main_sallary + total_add) - total_cut)

            # appending data to xlsx file
            page.append(finall_quiry)
        wb.save(filename = workbook_name)
        win10_display_notification(
            "رسالة تأكيد",
            f"تم انشاء التقرير وهو الان موجود في\n{workbook_name}\n"
        )


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = ApplicationUI()
    app.exec_()
